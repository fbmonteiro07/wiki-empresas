"""
build_ai_dashboards.py
----------------------
Réplica dos dashboards "Oficiais" da Fernanda (economics_1gw.html + como_flui_capacidade.html)
COM BASE NOS DADOS do "AI Model (comentado 2026-07-06).xlsx" (cópia do Felipe).

Faz o mesmo que atualizar_dashboards.py + atualizar_rcloud.py (os dois updaters da Fernanda),
mas:
  • lê com openpyxl (valores em cache) em vez de xlwings — não precisa do Excel aberto
    (xlwings não está instalado nesta máquina; pywin32/openpyxl estão);
  • lê do workbook COMENTADO em vez do "AI Model.xlsx" canônico;
  • escreve as réplicas em _wiki/_dashboards em vez de sobrescrever os originais da Fernanda;
  • conserta os links de navegação (../ai_model.html etc. não existem no destino).

Consts por dashboard:
  economics_1gw.html        : NET_PGW OTH_PGW CHIP_PGW CHIPS_PGW DES_GW DES_REV
  como_flui_capacidade.html : HYPER DES_GW DES_REV DES_GW_CLIENT DES_REV_CLIENT
                              AVGGW DESMIX LABSPLIT FLUXTAB

Uso:  py "E:\\Wiki Felipe empresas\\_wiki\\_tools\\build_ai_dashboards.py"
Só stdlib + openpyxl (sem pip).
"""

import os
import re
import math
import openpyxl

# ─── caminhos ───────────────────────────────────────────────────────────────
WB_PATH  = r"P:\Felipe Monteiro\US Equities\Modelos oficiais\AI Model (comentado 2026-07-06).xlsx"
SRC_DIR  = r"P:\Fernanda Neves\MarketData_PYTHON_CODES\Dashboards\Oficiais"
OUT_DIR  = r"E:\Wiki Felipe empresas\_wiki\_dashboards"
SRC_LABEL = "AI Model (comentado 2026-07-06).xlsx · uso interno · réplica"

# ═══════════════════════════════════════════════════════════════════════════════
# Shim xlwings→openpyxl: replica a semântica de sheet.range(...).value usada pelos
# pipelines originais (célula → escalar; linha/coluna → lista plana). Assim o código
# de leitura abaixo fica ~idêntico ao das fontes (atualizar_dashboards/atualizar_rcloud).
# ═══════════════════════════════════════════════════════════════════════════════
class _Val:
    def __init__(self, cells, single):
        self._cells, self._single = cells, single
    @property
    def value(self):
        return self._cells[0] if self._single else list(self._cells)

class Sheet:
    def __init__(self, ws):
        self.ws = ws
    def range(self, a, b=None):
        ws = self.ws
        if b is not None:                                   # range((r1,c1),(r2,c2))
            (r1, c1), (r2, c2) = a, b
            cells = [ws.cell(row=r, column=c).value
                     for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)]
            return _Val(cells, r1 == r2 and c1 == c2)
        if isinstance(a, tuple):                            # range((r,c))
            r, c = a
            return _Val([ws.cell(row=r, column=c).value], True)
        a = a.strip()                                       # range("A1") / range("A1:B1")
        if ":" in a:
            cells = [cell.value for row in ws[a] for cell in row]
            return _Val(cells, False)
        return _Val([ws[a].value], True)

# ─── helpers (copiados verbatim dos pipelines) ────────────────────────────────
def safe(val, dec=None):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return 0
    return round(val, dec) if dec is not None else val

def js_list(vals, dec=0):
    if dec == 0:
        return "[" + ",".join(str(round(v)) for v in vals) + "]"
    return "[" + ",".join(str(round(v, dec)) for v in vals) + "]"

def replace_const(html, name, new_val):
    """const NAME = [..]|{..}; — o (;) final força o não-guloso até o '}' do fecho."""
    pattern = rf"(const\s+{re.escape(name)}\s*=\s*)(\[.*?\]|\{{[\s\S]*?\}})(;)"
    new_html, n = re.subn(pattern, rf"\g<1>{new_val}\g<3>", html, count=1, flags=re.DOTALL)
    if n == 0:
        print(f"  [AVISO] const {name} não encontrada no HTML.")
    return new_html

def patch_obj(html, name, js):
    """const de OBJETO aninhado — casa até '\\n};'."""
    new, n = re.subn(rf"(const\s+{name}\s*=\s*)\{{[\s\S]*?\n\}};",
                     lambda m: m.group(1) + js + ";", html, count=1)
    if n == 0:
        print(f"  [AVISO] const {name} (obj) não encontrada no HTML.")
    return new

def jlist_gw(vals, d=2):
    return "[" + ",".join("null" if v is None else str(round(v, d)) for v in vals) + "]"

def jlist_int(vals):
    return "[" + ",".join("null" if v is None else str(round(v)) for v in vals) + "]"

# ═══════════════════════════════════════════════════════════════════════════════
print(f"Abrindo workbook comentado…\n  {WB_PATH}")
wb = openpyxl.load_workbook(WB_PATH, data_only=True)   # data_only=valores em cache
rt  = Sheet(wb["Resumo |  Total"])
bu  = Sheet(wb[" Bottom-up Capex"])
csp = Sheet(wb["Resumo CSPs"])

# ═══════════════════════════════════════════════════════════════════════════════
# 1. AI MODEL  (atualizar_dashboards.py) — cols CY24..CY28E: Resumo|Total D:H · BU Z:AD
# ═══════════════════════════════════════════════════════════════════════════════
def rt_row(row):
    return [safe(v) for v in rt.range(f"D{row}:H{row}").value]

def rt_label(r):
    for c in ("A", "B", "C"):
        v = rt.range(f"{c}{r}").value
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None

def _norm(s):
    return "" if not s else str(s).upper().split("(")[0].replace(" ", "").strip()

def find_header(substr, occ=1, lo=1, hi=200):
    cnt = 0
    for r in range(lo, hi + 1):
        l = rt_label(r)
        if l and substr.lower() in l.lower():
            cnt += 1
            if cnt == occ:
                return r
    return None

def csp_row(hdr, cspx, span=14):
    if hdr is None:
        return None
    tgt = _norm(cspx)
    for r in range(hdr + 1, hdr + span + 1):
        if _norm(rt_label(r)) == tgt:
            return r
    return None

H_INST = find_header("Estoque GW AI, EOP", 1)
H_ONL  = find_header("Estoque GW AI, EOP", 2)
H_AVG  = find_header("Estoque GW AI, AVG", 1)
H_CAPM = find_header("Capex Hyperscalers | Bottom-Up", 1)
H_CAPO = find_header("Capex OUTROS Players", 1)

MAJ = ["MSFT", "AMZN", "GOOGLE", "META"]
SUPPLY = MAJ + ["ORCL", "CRWV", "SpaceX"]

def _resid(total, parts):
    return [max(0.0, total[i] - sum(p[i] for p in parts)) for i in range(len(total))]

def block_sum(hdr, csps):
    tot = [0.0] * 5
    for cspx in csps:
        r = csp_row(hdr, cspx)
        if r:
            v = rt_row(r)
            for i in range(5):
                tot[i] += v[i]
    return tot

def hyper_map(hdr, csps):
    out = {}
    for cspx in csps:
        r = csp_row(hdr, cspx)
        out[cspx] = rt_row(r) if r else [0, 0, 0, 0, 0]
        if r is None:
            print(f"  [AVISO] '{cspx}' não achado no bloco linha {hdr}.")
    return out

def hyper_split(hdr, total_row, extra=()):
    m = hyper_map(hdr, SUPPLY + list(extra))
    m["OUTROS"] = _resid(total_row, [m[k] for k in SUPPLY + list(extra)])
    return m

gwai_map     = hyper_split(H_INST, rt_row(H_INST))
gwonline_map = hyper_split(H_ONL,  rt_row(H_ONL))
avgGw_map    = hyper_split(H_AVG,  rt_row(H_AVG), extra=["OAI", "ANTHROPIC"])

capex_map = hyper_map(H_CAPM, MAJ)
_capo = hyper_map(H_CAPO, ["ORCL", "CRWV"])
capex_map["ORCL"]   = _capo["ORCL"]
capex_map["CRWV"]   = _capo["CRWV"]
capex_map["SpaceX"] = [0, 0, 0, 0, 0]
capex_map["OUTROS"] = _resid(block_sum(H_CAPO, ["ORCL", "CRWV", "OAI", "ANTHROPIC", "OUTROS"]),
                             [capex_map["ORCL"], capex_map["CRWV"]])
print("  HYPER lido.")

def bu_row(row):
    return [safe(v) for v in bu.range(f"Z{row}:AD{row}").value]

def cumsum(vals):
    out, acc = [], 0.0
    for v in vals:
        acc += max(0.0, v)
        out.append(round(acc, 2))
    return out

DES_REV_ROWS = {"NVDA":134,"AVGO":135,"Google":136,"Meta":137,"OpenAI":138,
                "MRVL":139,"Trainium":140,"MSFT":141,"AMD":142,"MTK":143,"Alchip":144}
CLI_REV_ROWS = {"NVDA":148,"AMD":149,"Google":151,"Amazon":152,"Meta":153,"MSFT":154,"OpenAI":155}
des_rev        = {k: [round(x) for x in bu_row(r)] for k, r in DES_REV_ROWS.items()}
des_rev_client = {k: [round(x) for x in bu_row(r)] for k, r in CLI_REV_ROWS.items()}

DES_GW_ROWS = {"NVDA":99,"AVGO":100,"Google":101,"Meta":102,"OpenAI":103,
               "MRVL":104,"Trainium":105,"MSFT":106,"AMD":107,"MTK":108,"Alchip":109}
CLI_GW_ROWS = {"NVDA":112,"AMD":113,"Google":115,"Amazon":116,"Meta":117,"MSFT":118,"OpenAI":119}
des_gw        = {k: cumsum(bu_row(r)) for k, r in DES_GW_ROWS.items()}
des_gw_client = {k: cumsum(bu_row(r)) for k, r in CLI_GW_ROWS.items()}
print("  DES_GW / DES_REV / CLIENT lidos.")

CHIP_PGW_ROWS  = {"NVDA":170,"AVGO":171,"Google":172,"Meta":173,"OpenAI":174,
                  "MRVL":175,"Trainium":176,"MSFT":177,"AMD":178,"MTK":179,"Alchip":180}
CHIPS_PGW_ROWS = {"NVDA":36,"AVGO":37,"Google":38,"Meta":39,"OpenAI":40,
                  "MRVL":41,"Trainium":42,"MSFT":43,"AMD":44,"MTK":45,"Alchip":46}
net_pgw   = [round(x) for x in bu_row(251)]
oth_pgw   = [round(x) for x in bu_row(363)]
chip_pgw  = {k: [round(x) for x in bu_row(r)] for k, r in CHIP_PGW_ROWS.items()}
chips_pgw = {k: [round(x) for x in bu_row(r)] for k, r in CHIPS_PGW_ROWS.items()}
print("  ECONOMICS/GW lido.")

# ── formatadores de JS (verbatim) ─────────────────────────────────────────────
def fmt_hyper():
    def sub(d, dec=1, nonneg=False):
        lines = []
        for k, vals in d.items():
            v = [max(0, round(x, dec)) if nonneg else round(x, dec) for x in vals]
            fmt = js_list(v, dec) if dec > 0 else js_list(v)
            lines.append(f"    {k}:  {fmt}")
        return "{\n" + ",\n".join(lines) + "\n  }"
    return ("{\n"
            f"  gwonline:{sub(gwonline_map, dec=1, nonneg=True)},\n"
            f"  gwai:{sub(gwai_map, dec=1)},\n"
            f"  avgGw:{sub(avgGw_map, dec=3)},\n"
            f"  capex:{sub(capex_map, dec=0)}\n"
            "}")

def fmt_des(d, dec=1):
    return "{\n" + ",\n".join(
        f"  {k}:    {js_list([round(x, dec) for x in vals], dec)}" for k, vals in d.items()) + "\n}"

def fmt_des_rev(d):
    return "{\n" + ",\n".join(
        f"  {k}:    {js_list([round(x) for x in vals])}" for k, vals in d.items()) + "\n}"

# ═══════════════════════════════════════════════════════════════════════════════
# 2. FLUXO  (atualizar_rcloud.py) — AVGGW · DESMIX · LABSPLIT · FLUXTAB
# ═══════════════════════════════════════════════════════════════════════════════
# ── AVGGW: avg GW total por player (Resumo|Total D:H) ──
def _rt_avg(row):
    v = rt.range((row, 4), (row, 8)).value
    v = v if isinstance(v, list) else [v]
    return [round(0 if (x is None or (isinstance(x, float) and math.isnan(x))) else x, 2) for x in v]
avggw = {"total": _rt_avg(88), "msft": _rt_avg(89), "amzn": _rt_avg(90), "goog": _rt_avg(91), "meta": _rt_avg(92)}
avggw_js = "{\n" + ",\n".join(f"  {k}:{jlist_gw(avggw[k])}" for k in ["total","msft","amzn","goog","meta"]) + "\n}"
print("  AVGGW lido.")

# ── DESMIX: fluxograma designer→player (compras acumuladas), merchant split via Resumo CSPs ──
def _dm_row(r):
    v = bu.range((r, 26), (r, 30)).value  # Z:AD
    return [0.0 if (x is None or (isinstance(x, float) and math.isnan(x))) else float(x) for x in v]
DM_DELTA = {"NVDA": _dm_row(99), "GoogleTPU": _dm_row(101), "MetaMTIA": _dm_row(102),
            "Trainium": _dm_row(105), "AMD": _dm_row(107)}
DM_CAPTIVE = {"GoogleTPU": "GOOGLE", "Trainium": "AMZN", "MetaMTIA": "META"}
DM_MAPS = {"NVDA": {"MSFT":[7], "GOOGLE":[8], "AMZN":[9], "META":[10],
                    "SPACEX":[11], "ORCL":[12], "CRWV":[13], "OUTROS":[14]},
           "AMD":  {"MSFT":[18], "GOOGLE":[19], "AMZN":[20], "META":[21], "OUTROS":[22,23,24]}}
DM_PLAYERS = ["MSFT","AMZN","GOOGLE","META","ORCL","CRWV","SPACEX","OUTROS"]
def _dm_fracs(chip, yi):
    col = 26 + yi
    out = {p: 0.0 for p in DM_PLAYERS}
    for b, rr in DM_MAPS[chip].items():
        for rw in rr:
            v = csp.range((rw, col)).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                out[b] += v
    return out
desmix = {c: {p: [0.0] * 5 for p in DM_PLAYERS} for c in DM_DELTA}
for c, dl in DM_DELTA.items():
    acc = {p: 0.0 for p in DM_PLAYERS}
    for yi in range(5):
        d = max(0.0, dl[yi])
        if c in DM_CAPTIVE:
            acc[DM_CAPTIVE[c]] += d
        else:
            fr = _dm_fracs(c, yi)
            tot = sum(fr.values()) or 1.0
            for p in DM_PLAYERS:
                acc[p] += d * fr[p] / tot
        for p in DM_PLAYERS:
            desmix[c][p][yi] = round(acc[p], 3)
desmix_js = "{\n" + ",\n".join(
    "  " + c + ":{" + ",".join(f"{p}:{jlist_gw(desmix[c][p], 3)}" for p in DM_PLAYERS) + "}"
    for c in DM_DELTA) + "\n}"
print("  DESMIX lido.")

# ── LABSPLIT: GW treino/inferência OpenAI/Anthropic ('Resumo | Labs', por rótulo col B) ──
rl = Sheet(wb["Resumo | Labs"])
_rl_colB = ["" if x is None else str(x).strip().lower() for x in rl.range((1, 2), (200, 2)).value]
def _rl_find(txt, start=0):
    for i in range(start, len(_rl_colB)):
        if _rl_colB[i].startswith(txt):
            return i + 1
    return None
def _rl_vals(r):
    v = rl.range((r, 34), (r, 38)).value  # AH:AL
    return [round(0.0 if (x is None or (isinstance(x, float) and math.isnan(x))) else x, 3) for x in v]
_estoque = _rl_find("estoque gw ai labs")
_porlab  = _rl_find("por lab:", _estoque or 0)
labsplit = {}
for lab in ("openai", "anthropic"):
    r_lab = _rl_find(lab, _porlab or 0)
    r_tr  = _rl_find("treinamento", r_lab) if r_lab else None
    r_in  = _rl_find("inferencia", r_lab) if r_lab else None
    if r_tr is None or r_in is None:
        print(f"  [AVISO] LABSPLIT.{lab}: trein/infer não achado.")
        continue
    labsplit[lab] = (_rl_vals(r_tr), _rl_vals(r_in))
labsplit_js = None
if labsplit:
    labsplit_js = "{\n" + ",\n".join(
        f"  {k}:{{trein:{jlist_gw(t, 3)},infer:{jlist_gw(i, 3)}}}" for k, (t, i) in labsplit.items()) + "\n}"
    print("  LABSPLIT lido.")

# ── FLUXTAB: fluxo 4 camadas, bloco 'Fluxo GW' da Resumo|Total (2ª 'Estoque GW AI, AVG') ──
_ft_colB = ["" if x is None else str(x).strip() for x in rt.range((1, 2), (230, 2)).value]
def _ft_norm(s):
    return s.upper().replace(" ", "").replace(".", "")
def _ft_vals(r):
    v = rt.range((r, 4), (r, 8)).value  # D:H
    return [round(0.0 if (x is None or (isinstance(x, float) and math.isnan(x))) else float(x), 3) for x in v]
fluxtab_js = None
_ft_hits = [i + 1 for i, s in enumerate(_ft_colB) if s.lower().startswith("estoque gw ai, avg")]
if len(_ft_hits) < 2:
    print("  [AVISO] FLUXTAB: bloco 'Fluxo GW' não achado.")
else:
    _ft0 = _ft_hits[1]
    _FT_CHILD = {
        "INFERENCIA/TREINAMENTOOAI": "oai",  "DEMANDAIMPLICITAOAI": "oai",
        "INFERENCIA/TREINAMENTOANTHROPIC": "ant", "DEMANDAIMPLICITAANTHROPIC": "ant",
        "GEMINI": "gemini", "OTHERAI": "otherai", "SUPERFICIE": "superficie",
        "ESTTREINAMENTO": "trein", "OUTRASDEMANDAS": "outras",
        "USOINTERNO": "interno", "TREINAMENTOXAI": "xai",
    }
    _FT_HDRS = [("MSFT","MSFT"),("AMZN","AMZN"),("GOOGLE","GOOGLE"),("META","META"),
                ("ORCL","ORCL"),("CRWV","CRWV"),("SPACEX","SPACEX"),("OUTROS","OUTROS")]
    _hrows, _cur = [], _ft0
    _ok = True
    for hdr, key in _FT_HDRS:
        r = None
        for i in range(_cur, min(_cur + 40, len(_ft_colB))):
            if _ft_norm(_ft_colB[i].split("(")[0]) == hdr:
                r = i + 1
                break
        if r is None:
            print(f"  [AVISO] FLUXTAB: header '{hdr}' não achado.")
            _ok = False
            break
        _hrows.append((key, r))
        _cur = r
    if _ok:
        fluxtab = {}
        for j, (key, hr) in enumerate(_hrows):
            end = _hrows[j + 1][1] - 1 if j + 1 < len(_hrows) else hr + 4
            d = {"total": _ft_vals(hr)}
            for r in range(hr + 1, end + 1):
                n = _ft_norm(_ft_colB[r - 1])
                for pref, field in _FT_CHILD.items():
                    if n.startswith(pref):
                        d[field] = _ft_vals(r)
                        break
            fluxtab[key] = d
        _uso0 = _hrows[-1][1]
        _FT_LABS = [("INFERENCIAOAI","infOai"),("INFERENCIAANTHROPIC","infAnt"),
                    ("APIOAI","apiOai"),("APIANTHROPIC","apiAnt"),("APIGEMINI","apiGemini"),
                    ("TREINAMENTOOAI","treinOai"),("TREINAMENTOANTHROPIC","treinAnt")]
        labs = {}
        for pref, field in _FT_LABS:
            found = False
            for i in range(_uso0, min(_uso0 + 30, len(_ft_colB))):
                if _ft_norm(_ft_colB[i]) == pref:
                    labs[field] = _ft_vals(i + 1)
                    found = True
                    break
            if not found:
                labs[field] = [0.0] * 5
        fluxtab["labs"] = labs
        def _ft_js(d):
            return "{" + ",".join(f"{f}:{jlist_gw(v, 3)}" for f, v in d.items()) + "}"
        fluxtab_js = "{\n" + ",\n".join(f"  {k}:{_ft_js(d)}" for k, d in fluxtab.items()) + "\n}"
        print("  FLUXTAB lido.")

wb.close()

# ═══════════════════════════════════════════════════════════════════════════════
# 3. PATCH + ESCRITA
# ═══════════════════════════════════════════════════════════════════════════════
def clean_links(html):
    """Remove links '../' quebrados no destino (nav some; corpo vira texto) + rótulo de fonte."""
    def _nav(m):
        nav = m.group(0)
        nav = re.sub(r'[ \t]*<a href="\.\./[^"]*"[^>]*>[\s\S]*?</a>\n?', '', nav)
        return nav
    html = re.sub(r'<nav class="tabs">[\s\S]*?</nav>', _nav, html, count=1)
    html = re.sub(r'<a href="\.\./[^"]*"[^>]*>([\s\S]*?)</a>', r'\1', html)
    html = html.replace("AI Model.xlsx · uso interno", SRC_LABEL)
    return html

def read_template(name):
    """Usa a réplica já gerada (auto-suficiente, re-patchável) ou semeia do original da Fernanda."""
    out = os.path.join(OUT_DIR, name)
    if os.path.exists(out):
        with open(out, "r", encoding="utf-8") as f:
            return f.read()
    with open(os.path.join(SRC_DIR, name), "r", encoding="utf-8") as f:
        return f.read()

os.makedirs(OUT_DIR, exist_ok=True)

# ── economics_1gw.html ──
print("\nGerando economics_1gw.html…")
h = read_template("economics_1gw.html")
h = replace_const(h, "NET_PGW",   js_list(net_pgw))
h = replace_const(h, "OTH_PGW",   js_list(oth_pgw))
h = replace_const(h, "CHIP_PGW",  fmt_des_rev(chip_pgw))
h = replace_const(h, "CHIPS_PGW", fmt_des_rev(chips_pgw))
h = replace_const(h, "DES_GW",    fmt_des(des_gw, dec=2))
h = replace_const(h, "DES_REV",   fmt_des_rev(des_rev))
h = clean_links(h)
with open(os.path.join(OUT_DIR, "economics_1gw.html"), "w", encoding="utf-8") as f:
    f.write(h)
print("  economics_1gw.html escrito.")

# ── como_flui_capacidade.html ──
print("\nGerando como_flui_capacidade.html…")
h = read_template("como_flui_capacidade.html")
# link de volta ao par (o original não tinha; economics já linka pra cá)
if 'href="economics_1gw.html"' not in h:
    h = h.replace('<nav class="tabs">',
                  '<nav class="tabs">\n  <a href="economics_1gw.html">Qual o economics de 1 GW?</a>', 1)
h = replace_const(h, "HYPER",          fmt_hyper())
h = replace_const(h, "DES_GW",         fmt_des(des_gw, dec=2))
h = replace_const(h, "DES_REV",        fmt_des_rev(des_rev))
h = replace_const(h, "DES_GW_CLIENT",  fmt_des(des_gw_client, dec=2))
h = replace_const(h, "DES_REV_CLIENT", fmt_des_rev(des_rev_client))
h = patch_obj(h, "AVGGW", avggw_js)
h = patch_obj(h, "DESMIX", desmix_js)
if labsplit_js:
    h = patch_obj(h, "LABSPLIT", labsplit_js)
if fluxtab_js:
    h = patch_obj(h, "FLUXTAB", fluxtab_js)
h = clean_links(h)
with open(os.path.join(OUT_DIR, "como_flui_capacidade.html"), "w", encoding="utf-8") as f:
    f.write(h)
print("  como_flui_capacidade.html escrito.")

print("\nConcluído. Réplicas em:", OUT_DIR)
