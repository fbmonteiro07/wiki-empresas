"""Build TXN bottoms-up quarterly revenue model. openpyxl. English only."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- styles ----
HARD  = PatternFill("solid", fgColor="C6EFCE")   # green
ANCH  = PatternFill("solid", fgColor="BDD7EE")   # blue
EST   = PatternFill("solid", fgColor="FFE699")   # amber
HDR   = PatternFill("solid", fgColor="1F4E78")
SUB   = PatternFill("solid", fgColor="D9E1F2")
white = Font(color="FFFFFF", bold=True)
bold  = Font(bold=True)
ital  = Font(italic=True, size=9, color="555555")
thin  = Side(style="thin", color="BFBFBF")
box   = Border(left=thin, right=thin, top=thin, bottom=thin)
money = '#,##0'
pct   = '0.0%'

def hdr(ws, row, cols, start=1):
    for j,c in enumerate(cols):
        cell = ws.cell(row=row, column=start+j, value=c)
        cell.fill = HDR; cell.font = white; cell.border = box
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

# ============ ASSUMPTIONS ============
ws = wb.active; ws.title = "Assumptions"
ws["A1"] = "TXN — Bottoms-Up Quarterly Revenue Model — Assumptions"; ws["A1"].font = Font(bold=True, size=13)
ws["A2"] = "Build date 2026-07-01 · consensus asof 2026-07-01 (BBG) · segment build is HARD quarterly disclosure; end-market is annual-only (ANCHOR)"
ws["A2"].font = ital
ws["A4"] = "Legend:"; ws["A4"].font = bold
ws["B4"] = "HARD (filed/guided + cite)"; ws["B4"].fill = HARD; ws["B4"].border = box
ws["C4"] = "PARTIAL / ANCHOR"; ws["C4"].fill = ANCH; ws["C4"].border = box
ws["D4"] = "ESTIMATE"; ws["D4"].fill = EST; ws["D4"].border = box

r = 6
hdr(ws, r, ["Input", "Value", "Tag", "Source / basis"])
rows = [
 ("SEGMENT ACTUALS ($m) — from filings", None, None, None),
 ("Q1FY26 Analog revenue", 3924, "HARD", "10-Q filed 2026-04-24, segment note"),
 ("Q1FY26 Embedded revenue", 723, "HARD", "10-Q 2026-04-24"),
 ("Q1FY26 Other revenue", 178, "HARD", "10-Q 2026-04-24"),
 ("Q1FY26 Total (tie-out target)", 4825, "HARD", "10-Q 2026-04-24; +19% YoY, +9% QoQ"),
 ("Q4FY25 Total (Dec-25)", 4423, "HARD", "FY25 10-K 2026-02-06 less 9-mo 10-Q (Q1+Q2+Q3)"),
 ("Q3FY25 Total (Sep-25)", 4742, "HARD", "10-Q filed 2025-10-23"),
 ("Q2FY25 Total (Jun-25)", 4448, "HARD", "10-Q filed 2025-07-29 (3-mo table)"),
 ("Q1FY25 Total (Mar-25)", 4069, "HARD", "10-Q filed 2025-04-24"),
 ("NEXT-QUARTER GUIDE — HARD", None, None, None),
 ("Q2FY26 revenue guide LOW", 5000, "HARD", "Q1FY26 call 2026-04-22 (Lizardi)"),
 ("Q2FY26 revenue guide HIGH", 5400, "HARD", "Q1FY26 call 2026-04-22"),
 ("Q2FY26 guide MIDPOINT", 5200, "HARD", "midpoint (5000+5400)/2 ; ~+7.8% QoQ"),
 ("CALIBRATED GROWTH SLOPES (from observed QoQ)", None, None, None),
 ("Analog up-quarter QoQ (Q2/Q3)", 0.075, "ESTIMATE", "Observed up-qtr avg +8.0% (Q2F25 +7.5, Q3F25 +8.0, Q1F26 +8.5); use 7.5% as base"),
 ("Analog Q3 QoQ (moderating)", 0.06, "ESTIMATE", "Q3F25 tot was +6.6% QoQ; cycle maturing"),
 ("Analog Q4 seasonal QoQ", -0.02, "ESTIMATE", "Q4F25 was -3.1%; milder now on 2H pricing + DC (Ilan, Jefferies July hike)"),
 ("Analog Q1 up-quarter QoQ", 0.08, "ESTIMATE", "Q1F26 was +8.5% QoQ"),
 ("Embedded up-quarter QoQ", 0.035, "ESTIMATE", "~+4.5% (up-quarter average, excluding seasonal Q4 dip); +3.5% forecast conservative vs trailing; lags Analog in recovery"),
 ("Embedded Q4 seasonal QoQ", -0.03, "ESTIMATE", "Q4F25 Emb -6.6%; softened for pricing"),
 ("END-MARKET ANCHOR (annual %, 10-K MD&A)", None, None, None),
 ("Industrial % of revenue", 0.33, "ANCHOR", "FY25 10-K 2026-02-06 (annual, rounded, not quarterly)"),
 ("Automotive % of revenue", 0.33, "ANCHOR", "FY25 10-K 2026-02-06"),
 ("Personal electronics %", 0.21, "ANCHOR", "FY25 10-K 2026-02-06"),
 ("Data center %", 0.09, "ANCHOR", "FY25 10-K; DC +~90% YoY Q1F26 (call)"),
 ("Comms equipment %", 0.03, "ANCHOR", "FY25 10-K 2026-02-06"),
 ("Calculators %", 0.01, "ANCHOR", "FY25 10-K 2026-02-06"),
]
r += 1
for name, val, tag, src in rows:
    if val is None and tag is None:
        c = ws.cell(row=r, column=1, value=name); c.font = bold; c.fill = SUB
        for j in range(2,5): ws.cell(row=r, column=j).fill = SUB
        r += 1; continue
    ws.cell(row=r, column=1, value=name).border = box
    vc = ws.cell(row=r, column=2, value=val); vc.border = box
    vc.number_format = pct if isinstance(val,float) and val < 1 else money
    tc = ws.cell(row=r, column=3, value=tag); tc.border = box
    tc.fill = {"HARD":HARD, "ANCHOR":ANCH, "PARTIAL":ANCH, "ESTIMATE":EST}[tag]
    ws.cell(row=r, column=4, value=src).border = box
    r += 1
ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12; ws.column_dimensions["D"].width = 62

# ============ MODEL ============
ws = wb.create_sheet("Model")
ws["A1"] = "TXN — Quarterly Revenue Build ($m) — Analog / Embedded / Other"; ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = "Cols B-F = ACTUAL (tie to filed total). Cols G-J = FORECAST. Q2FY26 anchored to guide midpoint; forward slopes calibrated to observed QoQ."; ws["A2"].font = ital

qcols = ["Q1FY25","Q2FY25","Q3FY25","Q4FY25","Q1FY26","Q2FY26E","Q3FY26E","Q4FY26E","Q1FY27E"]
hdr(ws, 4, ["Line ($m)"] + qcols)
# actual segment data
# Q1F25,Q2F25,Q3F25,Q4F25,Q1F26 actual ; forecast built by formula
analog_act = [3210,3452,3729,3615,3924]
emb_act    = [647,679,709,662,723]
oth_act    = [212,317,304,146,178]

# Row 5 Analog, Row6 Embedded, Row7 Other, Row8 Total
ws.cell(row=5, column=1, value="Analog").font = bold
ws.cell(row=6, column=1, value="Embedded Processing").font = bold
ws.cell(row=7, column=1, value="Other").font = bold
ws.cell(row=8, column=1, value="TOTAL REVENUE").font = bold
for i,v in enumerate(analog_act):
    ws.cell(row=5, column=2+i, value=v).number_format = money
    ws.cell(row=6, column=2+i, value=emb_act[i]).number_format = money
    ws.cell(row=7, column=2+i, value=oth_act[i]).number_format = money
# Forecast columns G(7)=Q2F26 ... J(10)=Q1F27 -> col idx 7,8,9,10
# Analog forecast formulas referencing prior col and Assumptions slopes
# Assumptions cells: build a small param block on Model sheet for auditability
ws["A20"] = "Forecast parameters (calibrated slopes)"; ws["A20"].font = bold
params = [("Analog Q2 QoQ",0.075),("Analog Q3 QoQ",0.06),("Analog Q4 QoQ",-0.02),("Analog Q1F27 QoQ",0.08),
          ("Emb Q2 QoQ",0.035),("Emb Q3 QoQ",0.03),("Emb Q4 QoQ",-0.03),("Emb Q1F27 QoQ",0.06)]
pr = 21
pcell = {}
for nm,vv in params:
    ws.cell(row=pr,column=1,value=nm).border=box
    c=ws.cell(row=pr,column=2,value=vv); c.number_format=pct; c.fill=EST; c.border=box
    pcell[nm] = f"$B${pr}"
    pr+=1
# Other forecast (seasonal estimate, hardcoded but tagged amber via note)
oth_fc = [234,250,150,185]  # Q2F26..Q1F27
# Analog forecast formulas
# col letters: F=Q1F26(6), G=Q2F26(7)...
def L(c): return get_column_letter(c)
# Analog
ws.cell(row=5, column=7, value=f"=ROUND({L(6)}5*(1+{pcell['Analog Q2 QoQ']}),0)").number_format=money
ws.cell(row=5, column=8, value=f"=ROUND({L(7)}5*(1+{pcell['Analog Q3 QoQ']}),0)").number_format=money
ws.cell(row=5, column=9, value=f"=ROUND({L(8)}5*(1+{pcell['Analog Q4 QoQ']}),0)").number_format=money
ws.cell(row=5, column=10, value=f"=ROUND({L(9)}5*(1+{pcell['Analog Q1F27 QoQ']}),0)").number_format=money
# Embedded
ws.cell(row=6, column=7, value=f"=ROUND({L(6)}6*(1+{pcell['Emb Q2 QoQ']}),0)").number_format=money
ws.cell(row=6, column=8, value=f"=ROUND({L(7)}6*(1+{pcell['Emb Q3 QoQ']}),0)").number_format=money
ws.cell(row=6, column=9, value=f"=ROUND({L(8)}6*(1+{pcell['Emb Q4 QoQ']}),0)").number_format=money
ws.cell(row=6, column=10, value=f"=ROUND({L(9)}6*(1+{pcell['Emb Q1F27 QoQ']}),0)").number_format=money
# Other (estimate)
for i,v in enumerate(oth_fc):
    c=ws.cell(row=7, column=7+i, value=v); c.number_format=money; c.fill=EST
# Totals (formula, all cols)
for c in range(2,11):
    tc=ws.cell(row=8, column=c, value=f"=SUM({L(c)}5:{L(c)}7)"); tc.number_format=money; tc.font=bold; tc.fill=SUB
# QoQ and YoY rows
ws.cell(row=10, column=1, value="Total QoQ %").font=ital
for c in range(3,11):
    ws.cell(row=10, column=c, value=f"={L(c)}8/{L(c-1)}8-1").number_format=pct
ws.cell(row=11, column=1, value="Total YoY %").font=ital
for c in range(6,11):
    ws.cell(row=11, column=c, value=f"={L(c)}8/{L(c-4)}8-1").number_format=pct
# tie-out check
ws.cell(row=13, column=1, value="TIE-OUT CHECK (Q1FY26)").font=bold
ws.cell(row=13, column=2, value="=F8").number_format=money
ws.cell(row=13, column=3, value=4825).number_format=money
ws.cell(row=13, column=4, value="=IF(F8=4825,\"PASS\",\"FAIL\")").font=bold
# FY totals
ws.cell(row=15, column=1, value="CY2026E (Q1F26 act + Q2-Q4F26E)").font=bold
ws.cell(row=15, column=2, value="=F8+G8+H8+I8").number_format=money
ws.cell(row=16, column=1, value="CY2026E QoQ note").font=ital
ws.cell(row=16, column=2, value="=SUM(F5:I5)").number_format=money  # analog CY26
for c in range(1,11): ws.column_dimensions[L(c)].width = 11
ws.column_dimensions["A"].width = 30

# ============ vs CONSENSUS ============
ws = wb.create_sheet("vs Consensus")
ws["A1"]="TXN — Bottoms-Up vs BBG Consensus & vs Guidance"; ws["A1"].font=Font(bold=True,size=12)
ws["A2"]="BBG BEST_SALES asof 2026-07-01. Nearest-quarter also checked vs TXN's own guided range (sharper than consensus)."; ws["A2"].font=ital
hdr(ws,4,["Period","Bottoms-up ($m)","BBG cons ($m)","$ var","% var","TXN guide range","Inside guide?"])
data=[
 ("Q2FY26E", "=Model!G8", 5211, "$5,000–5,400", "Q2 guide (call 2026-04-22)"),
 ("Q3FY26E", "=Model!H8", 5570, "n/a (not guided)", "beyond guided qtr"),
 ("Q4FY26E", "=Model!I8", None, "n/a", "no discrete cons qtr"),
 ("CY2026E", "=Model!F8+Model!G8+Model!H8+Model!I8", 20689, "n/a", "vs cons FY"),
 ("CY2027E", 24606, 23421, "n/a", "vs cons FY (extended build)"),
]
rr=5
for per,bu,cons,grange,note in data:
    ws.cell(row=rr,column=1,value=per).border=box
    bc=ws.cell(row=rr,column=2,value=bu); bc.number_format=money; bc.border=box
    if cons:
        cc=ws.cell(row=rr,column=3,value=cons); cc.number_format=money; cc.border=box
        ws.cell(row=rr,column=4,value=f"=B{rr}-C{rr}").number_format=money
        ws.cell(row=rr,column=4).border=box
        ws.cell(row=rr,column=5,value=f"=B{rr}/C{rr}-1").number_format=pct
        ws.cell(row=rr,column=5).border=box
    else:
        for cc in (3,4,5): ws.cell(row=rr,column=cc,value="n/a").border=box
    ws.cell(row=rr,column=6,value=grange).border=box
    ws.cell(row=rr,column=7,value=note).border=box
    rr+=1
# explicit inside-guide check for Q2
ws.cell(row=12,column=1,value="Q2FY26 inside guide 5,000–5,400?").font=bold
ws.cell(row=12,column=2,value="=IF(AND(Model!G8>=5000,Model!G8<=5400),\"YES\",\"NO\")").font=bold
# CY2027 confidence flag (visible, load-bearing extrapolation warning)
warn_fill = PatternFill("solid", fgColor="FFC000")  # orange
wc = ws.cell(row=9, column=6, value="[!] 7 of 8 quarters extrapolated (0 explicitly guided) — treat CY27 as directional, not precise")
wc.fill = warn_fill; wc.font = Font(bold=True, size=9); wc.border = box
wc.alignment = Alignment(wrap_text=True, vertical="center")
ws.merge_cells(start_row=9, start_column=6, end_row=9, end_column=7)
ws.cell(row=9, column=7).border = box
ws.row_dimensions[9].height = 42
# also a standalone banner below the table for unmissable visibility
bc = ws.cell(row=14, column=1, value="[!] CY2027E CONFIDENCE FLAG: only Q2FY26 is hard-guided ($5.0-5.4B). 7 of the 8 quarters in the CY27 comparison are EXTRAPOLATED off the observed up-quarter recovery slope (Analog ~+7.5-8%/qtr). The +5.1% above-Street headline holds only if that slope persists for all 7 unguided quarters. Treat CY27 as directional, not precise. See Sensitivity tab flat-sequential row for the downside stress case.")
bc.fill = warn_fill; bc.font = Font(bold=True, size=10); bc.border = box
bc.alignment = Alignment(wrap_text=True, vertical="center")
ws.merge_cells(start_row=14, start_column=1, end_row=14, end_column=7)
for cc in range(1,8): ws.cell(row=14, column=cc).border = box
ws.row_dimensions[14].height = 70
for c,w in zip("ABCDEFG",[16,16,14,10,9,18,26]): ws.column_dimensions[c].width=w

# ============ SENSITIVITY ============
ws = wb.create_sheet("Sensitivity")
ws["A1"]="TXN — Sensitivity: load-bearing ESTIMATE inputs at +/-25%"; ws["A1"].font=Font(bold=True,size=12)
ws["A2"]="Two most sensitive: (1) Analog recovery slope (industrial), (2) Analog Q4 seasonal dip (auto/China + 2H pricing offset)."; ws["A2"].font=ital
hdr(ws,4,["Scenario","Analog Q2/Q3/Q1F27 slope","CY26E rev ($m)","CY27E rev ($m)","vs base CY26","vs base CY27"])
# base slopes: Q2 .075, Q3 .06, Q1F27 .08 ; sensitivity scales these +/-25%
# recompute in python for display (static values), base already known
def build(sl_mult, q4_qoq=-0.02):
    a=3924
    # Q2..Q4 F26, Q1F27
    q2=(4218 if sl_mult==1 else round(a*(1+0.075*sl_mult)))
    a=q2; q3=round(a*(1+0.06*sl_mult)); a=q3
    q4=round(a*(1+q4_qoq)); a=q4
    q1f27=round(a*(1+0.08*sl_mult))
    # embedded + other held at base
    e=[748,771,748,792]; o=[234,250,150,185]
    cy26=4825 + (q2+e[0]+o[0]) + (q3+e[1]+o[1]) + (q4+e[2]+o[2])
    # FY27 extend
    aa=q1f27
    r2=round(aa*1.075); aa=r2; r3=round(aa*1.06); aa=r3; r4=round(aa*0.98)
    ee=[820,845,820]; oo=[245,250,150]
    cy27=(q1f27+792+185)+(r2+ee[0]+oo[0])+(r3+ee[1]+oo[1])+(r4+ee[2]+oo[2])
    return cy26,cy27
base26,base27=build(1.0)
scen=[("Slope +25%",1.25),("Base",1.0),("Slope -25%",0.75)]
rr=5
for nm,m in scen:
    c26,c27=build(m)
    ws.cell(row=rr,column=1,value=nm).border=box
    ws.cell(row=rr,column=2,value=f"{0.075*m:.3f}/{0.06*m:.3f}/{0.08*m:.3f}").border=box
    ws.cell(row=rr,column=3,value=c26).number_format=money; ws.cell(row=rr,column=3).border=box
    ws.cell(row=rr,column=4,value=c27).number_format=money; ws.cell(row=rr,column=4).border=box
    ws.cell(row=rr,column=5,value=c26-base26).number_format=money; ws.cell(row=rr,column=5).border=box
    ws.cell(row=rr,column=6,value=c27-base27).number_format=money; ws.cell(row=rr,column=6).border=box
    rr+=1
# Q4 seasonal sensitivity
ws.cell(row=10,column=1,value="Q4 seasonal-dip sensitivity (Analog Q4 QoQ)").font=bold
hdr(ws,11,["Scenario","Analog Q4 QoQ","CY26E rev ($m)","CY27E rev ($m)","vs base CY26","vs base CY27"])
for nm,q4 in [("Mild dip -1.5%",-0.015),("Base -2.0%",-0.02),("Deep dip -6.7% (=Q4F25)",-0.067)]:
    c26,c27=build(1.0,q4)
    rr=12 if nm.startswith("Mild") else (13 if nm.startswith("Base") else 14)
    ws.cell(row=rr,column=1,value=nm).border=box
    ws.cell(row=rr,column=2,value=q4).number_format=pct; ws.cell(row=rr,column=2).border=box
    ws.cell(row=rr,column=3,value=c26).number_format=money; ws.cell(row=rr,column=3).border=box
    ws.cell(row=rr,column=4,value=c27).number_format=money; ws.cell(row=rr,column=4).border=box
    ws.cell(row=rr,column=5,value=c26-base26).number_format=money; ws.cell(row=rr,column=5).border=box
    ws.cell(row=rr,column=6,value=c27-base27).number_format=money; ws.cell(row=rr,column=6).border=box
# ---- NORMALIZATION / FLAT-SEQUENTIAL SCENARIO (skeptic's downside stress test) ----
def build_flat():
    # Q2FY26 keeps the guided recovery (only hard-guided qtr); Analog & Embedded held
    # FLAT (0% QoQ) from Q3FY26E through end of forecast window. No further cycle recovery.
    q2=4218                 # Q2F26 analog (guided, = base)
    q3=q2; q4=q3; q1f27=q4  # flat sequential
    eQ2=748; ef=eQ2         # Q2F26 emb (= base), then flat
    o=[234,250,150,185]     # other Q2..Q1F27 (= base tab)
    cy26=4825 + (q2+ef+o[0]) + (q3+ef+o[1]) + (q4+ef+o[2])
    # FY27 all flat: analog r2=r3=r4=q1f27 ; emb flat ; other = base tab FY27 other
    r2=q1f27; r3=q1f27; r4=q1f27
    oo=[245,250,150]
    cy27=(q1f27+ef+185)+(r2+ef+oo[0])+(r3+ef+oo[1])+(r4+ef+oo[2])
    return cy26,cy27
flat26,flat27=build_flat()
ws.cell(row=16,column=1,value="NORMALIZATION SCENARIO (skeptic downside stress)").font=bold
hdr(ws,17,["Scenario","Analog/Emb QoQ from Q3FY26E","CY26E rev ($m)","CY27E rev ($m)","vs base CY27","% vs BBG cons CY27"])
BBG_CY27=23421
row=18
ws.cell(row=row,column=1,value="Flat-sequential from Q3FY26E onward (no further recovery, cycle normalizes)").border=box
ws.cell(row=row,column=1).alignment=Alignment(wrap_text=True,vertical="center")
ws.cell(row=row,column=2,value="0.0% / 0.0% (held flat)").border=box
ws.cell(row=row,column=3,value=flat26).number_format=money; ws.cell(row=row,column=3).border=box
ws.cell(row=row,column=4,value=flat27).number_format=money; ws.cell(row=row,column=4).border=box
ws.cell(row=row,column=5,value=flat27-base27).number_format=money; ws.cell(row=row,column=5).border=box
vc=ws.cell(row=row,column=6,value=flat27/BBG_CY27-1); vc.number_format=pct; vc.border=box
ws.cell(row=row,column=4).fill=PatternFill("solid",fgColor="F8CBAD")  # highlight downside
ws.cell(row=row,column=6).fill=PatternFill("solid",fgColor="F8CBAD")
ws.row_dimensions[row].height=42
_pct27 = (flat27/BBG_CY27-1)*100
ws.cell(row=20,column=1,value=("Base case = +5.1% ABOVE Street on CY27. Flat-sequential collapses CY27 to ${f27} ({p27:.1f}% vs BBG cons ${b27}) — i.e. ~10-12% BELOW consensus. "
    "This is the swing from the recovery-slope assumption: base case assumes Analog ~+7.5-8%/qtr persists through 7 unguided quarters; flat case assumes the cycle stops recovering after the one guided quarter (Q2FY26). Directional only.").format(f27=format(flat27,",d"), p27=_pct27, b27=format(BBG_CY27,",d")))
ws.cell(row=20,column=1).font=ital
ws.cell(row=20,column=1).alignment=Alignment(wrap_text=True,vertical="top")
ws.merge_cells(start_row=20,start_column=1,end_row=20,end_column=6)
ws.row_dimensions[20].height=58
for c,w in zip("ABCDEF",[26,22,14,14,12,16]): ws.column_dimensions[c].width=w

# ============ SOURCES ============
ws = wb.create_sheet("Sources")
ws["A1"]="TXN Model — Sources"; ws["A1"].font=Font(bold=True,size=12)
hdr(ws,3,["Type","Document","Date","Used for"])
srcs=[
 ("Filing","TXN 10-K FY2025 (acc 0000097476-26-000059)","2026-02-06","FY25/24/23 annual segment rev; end-market % (annual)"),
 ("Filing","TXN 10-Q Q1FY26 (acc ...-26-000101)","2026-04-24","Q1FY26 segment rev (tie-out); Q1FY25 comp"),
 ("Filing","TXN 10-Q Q3FY25 (acc ...-25-000060)","2025-10-23","Q3FY25 quarter + 9-mo (Q4 back-out)"),
 ("Filing","TXN 10-Q Q2FY25 (acc ...-25-000036)","2025-07-29","Q2FY25 quarter (3-mo table)"),
 ("Filing","TXN 10-Q Q1FY25 (acc ...-25-000027)","2025-04-24","Q1FY25 quarter"),
 ("Transcript","TXN Q1FY26 earnings call","2026-04-22","Q2FY26 guide $5.0-5.4bn; end-market color; capex/FCF"),
 ("Transcript","TXN Q4FY25 earnings call","2026-01-27","Q1FY26 guide; DC +70% YoY; loadings/GM"),
 ("Transcript","TXN Q3FY25 earnings call","2025-10-22","Industrial recovery pace; DC $1.2bn run-rate"),
 ("Equity call","UBS/Arcuri TXN callback","2026-04-23","Pricing scrapes (TXN sandbagging 2H); industrial pull-in debate"),
 ("Equity call","BofA/Arya TXN analog","2026-04-23","AI/DC exposure framing; own-analog H1 call"),
 ("Consensus","BBG estimates.json companies.TXN (asof)","2026-07-01","BEST_SALES Q2/Q3/CY26/CY27; px 298.41"),
 ("Wiki","_wiki/TXN.md (thesis/debate)","2026-06-19","Broker PTs/stances; end-market mix; DC bull/bear"),
]
rr=4
for t,d,dt,u in srcs:
    for j,v in enumerate([t,d,dt,u]):
        ws.cell(row=rr,column=1+j,value=v).border=box
    rr+=1
for c,w in zip("ABCD",[12,48,12,58]): ws.column_dimensions[c].width=w

out=r"E:\Wiki Felipe empresas\_wiki\_models\TXN_bottom_up_model.xlsx"
wb.save(out)
print("SAVED", out)
print("base CY26", base26, "base CY27", base27)
